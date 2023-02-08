# pySimpleGUI Version
import sys

import PySimpleGUI as sg
from PySimpleGUI.PySimpleGUI import R

import numpy

value = sg.popup_get_file('TOMASでダウンロードした入力チェックファイルを選択してください。')#使用するダウンロード済みの勤務表元ファイルを選択

#nt_kinmu = sg.popup_get_file('休暇パターンで作成したファイルを選択してください。')#使用するダウンロード済みの勤務表元ファイルを選択

sg.theme('Python')

layout =[[sg.Text('THOMAS・[支社事業所委託業務管理]',font = ('Noto Serif CJK JP',10))],
        [sg.Text('[委託実施報告　入力データチェック出力] で出力したファイルを準備',font = ('meiryo',10))],
        [sg.Text('年月を入力',text_color='#FF0000',font =( 'meiryo,8')),sg.InputText(size = (10,2),key= '-YM-')],
        # [sg.Text('月前半・後半',text_color='#FF0000',font =( 'meiryo,8')),sg.Combo(['月前半','月後半'],size = (10,2),key ='-UPLW-')],
        [sg.Button('入力', button_color=('red','#808080'),key = '-SUBMIT-'),sg.Text('入力ボタンを押した後,Windowを閉じてください。',font = ('Noto Serif CJK JP',10))]]

window = sg.Window('TOMAS勤務入力チェックAPP',layout,size = (500,150))

while True:
   event,values = window.read()
   if event == '-SUBMIT-':

            num = values['-YM-']
            print(num)
       


   if event == sg.WIN_CLOSED:
         break

window.close()

print(num)


#-------------------------------------------------------------------


# from os import X_OK
import openpyxl
from openpyxl.styles import PatternFill

print('取得した値')

print(num)
# print(num1)

# yomikomisaki = 'c:/Users/406239/OneDrive - (株)NHKテクノロジーズ/デスクトップ/Tomas 入力チェックプログラム/{0}_5512沖縄管内_沖縄事業所_委託入力チェック.xlsx'.format(num)


#↑沖縄管内_沖縄事業所_委託入力チェック　　沖縄管内事業所_委託入力チェックに変更。20220715

#yomikomisaki = '/Users/gongyi1/Desktop/python/{0}_5512沖縄管内_沖縄事業所_委託入力チェック.xlsx'.format(num)

# GUIで入力した値を読み込みファイル名に代入させる

import pandas as pd

#df_use_skip=pd.read_excel(yomikomisaki,sheet_name='5512',header=0,index_col=None,usecols=[4,5,3],skiprows=[-1],skipfooter=0)

df_master=pd.read_excel(value,sheet_name='5512',header=0,index_col=None,skiprows=[-1],skipfooter=0)

print('[取り込んだ全データ表示]')
2
print(df_master)  #取得全データー表示

#追加時間が欠測値の行を削除（休日など実働無しの日）

# df_master.dropna(subset = ['追加時間'],inplace = True)

df_master['確認メッセージ'].mask((df_master['追加時間'] != '00:00') & (df_master['実施業務内容'].isnull() == True) & (df_master['追加時間'].isnull() != True) , '内容未記入',inplace = True)



print('[取り込んだ全データ表示2]')

print(df_master)  #取得全データー表示

df_master.to_excel('{0}_TomasCheck_pvw.xlsx'.format(num))


filename = 'c:/Users/406239/OneDrive - (株)NHKテクノロジーズ/デスクトップ/Tomas 入力チェックプログラム/{0}_TomasCheck_pvw.xlsx'.format(num)
wb = openpyxl.load_workbook(filename)
ws = wb['Sheet1']    

lastrow = ws.max_row
lastcol = ws.max_column

for col in range(2,lastrow +1):

   for row in range(2,lastrow+1):
        val = ws.cell(row= row,column = col).value

        if val == '内容未記入':
            ws.cell(row = row,column=col).fill= PatternFill(fill_type ='solid',fgColor = 'FFFF00')

df_master.to_excel('{0}_TomasCheck_pvw.xlsx'.format(num))

wb.save('{0}_TomasCheck.xlsx'.format(num))#印刷用ファイルに書き込み（最終版）


    
